from django.shortcuts import render, redirect
from .forms import *
from django.contrib import messages
from django.http import HttpResponse
from openpyxl import Workbook
import openpyxl
from django.core.exceptions import PermissionDenied


def lop_giang_day(request):
    if request.user.is_superuser or not request.user.is_staff:
        raise PermissionDenied

    user = User.objects.get(username=request.user)
    gv = GiaoVien.objects.get(user=user)
    ds_lop = Lop.objects.filter(giang_vien=gv)
    for i in ds_lop:
        i.count = SinhVien.objects.filter(lop=i).count()

    a = request.POST.get('id_lop')
    request.session['lop'] = request.POST.get('ds_mon_day')

    context = {
        'ds_lop': ds_lop,
        'id_lop': request.session['lop']
    }
    return render(request, 'lop_giang_day.html', context)


def nhap_diem(request, id):
    if not request.user.is_staff:
        raise PermissionDenied
    lop = Lop.objects.get(pk=id)
    request.session['id_nhap_diem'] = lop.id
    ds_sv = SinhVien.objects.filter(lop=lop)
    for sv in ds_sv:
        try:
            sv.diem = Diem.objects.get(
                sinh_vien=sv, mon_hoc__ma_mon=lop.mon_hoc.ma_mon)
            new_diem = sv.diem.diem_chuyen_can * 0.1 + \
                sv.diem.diem_kiem_tra * 0.2 + sv.diem.diem_cuoi_ki * 0.7
            sv.diem.tong_ket = round(new_diem, 2)
            if sv.diem.tong_ket < 4.0 or sv.diem.diem_chuyen_can <= 2:
                sv.diem.diem_chu = 'F'
            elif sv.diem.tong_ket >= 4.0 and sv.diem.tong_ket < 5:
                sv.diem.diem_chu = 'D'
            elif sv.diem.tong_ket >= 5.0 and sv.diem.tong_ket < 5.5:
                sv.diem.diem_chu = 'D+'
            elif sv.diem.tong_ket >= 5.5 and sv.diem.tong_ket < 6.5:
                sv.diem.diem_chu = 'C'
            elif sv.diem.tong_ket >= 6.5 and sv.diem.tong_ket < 7:
                sv.diem.diem_chu = 'C+'
            elif sv.diem.tong_ket >= 7.0 and sv.diem.tong_ket < 8:
                sv.diem.diem_chu = 'B'
            elif sv.diem.tong_ket >= 8.0 and sv.diem.tong_ket < 8.5:
                sv.diem.diem_chu = 'B+'
            elif sv.diem.tong_ket >= 8.5 and sv.diem.tong_ket <= 10:
                sv.diem.diem_chu = 'A'

            if sv.diem.tong_ket < 4.0 or sv.diem.diem_chuyen_can <= 2:
                sv.diem.danh_gia = 'HOCLAI'
            elif sv.diem.tong_ket >= 4.0:
                sv.diem.danh_gia = 'DAT'
        except:
            sv.diem = ''
    new_session = request.session.get('select', None)
    print(ds_sv)
    for sv in ds_sv:
        try:
            sv.item = Diem.objects.get(sinh_vien=sv, mon_hoc=lop.mon_hoc)
            print(sv)
        except:
            pass
    context = {
        'lop': lop,
        'diem_tp': ds_sv
    }
    return render(request, 'nhap_diem.html', context)


def insert_data(request):
    if request.user.is_superuser or not request.user.is_staff:
        raise PermissionDenied

    receiver = request.session['id_nhap_diem']
    ds_sv = SinhVien.objects.filter(lop=receiver)
    l = Lop.objects.get(id=receiver)
    ok = True
    for sv in ds_sv:
        diem_cc = request.POST.get('diem_cc_' + str(sv.id))
        diem_kt = request.POST.get('diem_kt_' + str(sv.id))
        diem_ck = request.POST.get('diem_ck_' + str(sv.id))

        if diem_cc == '' or diem_kt == '' or diem_ck == '':
            ok = False
        else:
            cc = float(diem_cc)
            kt = float(diem_kt)
            ck = float(diem_ck)
            try:
                d = Diem.objects.get(
                    sinh_vien=sv,
                    mon_hoc=l.mon_hoc
                )
                d.diem_chuyen_can = cc
                d.diem_kiem_tra = kt
                d.diem_cuoi_ki = ck
                d.save()
            except:
                Diem.objects.create(
                    sinh_vien=sv, mon_hoc=l.mon_hoc,
                    diem_chuyen_can=cc,
                    diem_kiem_tra=kt,
                    diem_cuoi_ki=ck
                )
            d = Diem.objects.filter(sinh_vien=sv, mon_hoc=l.mon_hoc)
    # Tạo lịch sử
    try:
        h = History.objects.get(
            user=request.user, actions=f'Đã cập nhật điểm lớp: {l}')
        h.save()
    except:
        History.objects.create(
            user=request.user,
            actions=f'Đã cập nhật điểm lớp: {l}'
        )

    if ok:
        messages.add_message(request, messages.SUCCESS,
                             'Cập nhật điểm thành công ! ')
    else:
        messages.add_message(request, messages.WARNING,
                             'Các trường không được bỏ trống !')
    return redirect('nhap_diem', receiver)


def export_csv(request):
    if request.user.is_superuser or not request.user.is_staff:
        raise PermissionDenied

    receiver = request.session['id_nhap_diem']
    ds_sv = SinhVien.objects.filter(lop=receiver)
    l = Lop.objects.get(id=receiver)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename={l}-diem.xlsx'.format(
        l=l.ma_lop
    )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Hello'

    columns = [
        'STT',
        'Mã SV',
        'Tên SV',
        'Chuyên cần',
        'Kiểm tra',
        'Cuối kì',
    ]
    row_num = 1
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for sv in ds_sv:
        row_num += 1

        # Define the data for each cell in the row
        try:
            diem = Diem.objects.get(
                sinh_vien=sv,
                mon_hoc=l.mon_hoc
            )
            row = [
                row_num - 1,
                sv.user.username,
                sv.hoten,
                diem.diem_chuyen_can,
                diem.diem_kiem_tra,
                diem.diem_cuoi_ki,
            ]

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
        except:
            break
    workbook.save(response)

    return response


def nhap_diem_excel(request):
    if request.user.is_superuser or not request.user.is_staff:
        raise PermissionDenied

    id_lop = request.session['id_nhap_diem']
    mon_hoc_id = Lop.objects.get(pk=id_lop).mon_hoc
    lop_name = Lop.objects.get(pk=id_lop)
    sv = SinhVien.objects.filter(lop__id=id_lop)
    if request.method == 'GET':
        return redirect('nhap_diem', id_lop)
    else:
        file_name = request.FILES.get('ex_file')
        wb = openpyxl.load_workbook(file_name)

        sheet = wb.active
        max_col = sheet.max_column
        max_row = sheet.max_row
        arr_all = []
        error_arr = []
        index_error = False
        index_error_text = ''
        ap = False
        commit = False

        fieldIndexes = {}

        for j in range(1, max_col+1):
            c = sheet.cell(row=1, column=j)
            fieldIndexes[str(c.value).strip()] = j-1

        for i in range(1, max_row+1):
            arr = []
            for j in range(1, max_col+1):
                c = sheet.cell(row=i, column=j)
                arr.append(c.value)
            arr_all.append(arr)
        for item in arr_all:
            try:
                if item[1] is None or item[2] is None or item[3] is None:
                    error_arr.append(item[0])
                    ap = True
            except IndexError as e:
                index_error_text = str(e)
                index_error = True
                break
            for m in sv:
                if str(m.user) == str(item[fieldIndexes["Mã SV"]]).strip():
                    print(item[fieldIndexes['Cuối kì']])
                    try:
                        q = Diem.objects.get(sinh_vien=m, mon_hoc=mon_hoc_id)
                        if Diem.objects.filter(sinh_vien=m, mon_hoc=mon_hoc_id).exists():
                            new_data = Diem.objects.get(
                                sinh_vien=m, mon_hoc=mon_hoc_id)
                            new_data.diem_chuyen_can = item[fieldIndexes['Chuyên cần']]
                            new_data.diem_kiem_tra = item[fieldIndexes['Kiểm tra']]
                            new_data.diem_cuoi_ki = item[fieldIndexes['Cuối kì']]
                            new_data.save()
                    except:
                        Diem.objects.create(
                            sinh_vien=m,
                            mon_hoc=mon_hoc_id,
                            diem_chuyen_can=item[fieldIndexes['Chuyên cần']],
                            diem_kiem_tra=item[fieldIndexes['Kiểm tra']],
                            diem_cuoi_ki=item[fieldIndexes['Cuối kì']]
                        )
                    break
                commit = True
        try:
            h = History.objects.get(
                user=request.user, actions=f'Đã cập nhật điểm lớp: {lop_name}')
            h.save()
        except:
            History.objects.create(
                user=request.user,
                actions=f'Đã cập nhật điểm lớp: {lop_name}'
            )
        if commit:
            messages.add_message(request, messages.SUCCESS,
                                 'Đã cập nhật.')
        if index_error:
            messages.add_message(
                request, messages.ERROR, f'Số lượng các trường trong file không hợp lệ - {index_error_text}')

    return redirect('nhap_diem', id_lop)
