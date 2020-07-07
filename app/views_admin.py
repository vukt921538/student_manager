from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login as auth_login
from . forms import *
import random
import datetime
from django.core.files.storage import FileSystemStorage
from django.contrib import messages
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from django.core.exceptions import ValidationError
import numpy as np
from django.core.exceptions import PermissionDenied


def create(request):
    if not request.user.is_superuser:
        raise PermissionDenied
    form = CreateStudent()
    if request.method == 'POST':
        nienkhoa = request.POST.get('nien_khoa')
        nganh = request.POST.get('nganh')
        hedt = request.POST.get('hedt')

        try:
            get_nganh = Nganh.objects.get(pk=nganh)
            get_hedt = HeDT.objects.get(pk=hedt)
            get_nienkhoa = NienKhoa.objects.get(pk=nienkhoa)

            # Trả về 1 array với 1 phần tử
            number = random.sample(range(1000), 1)
            # => Convert về string
            # Sử dụng 2 phương thữ xử lý chuỗi : thay thế [ = '' và xoá phần tử ]
            new_string = str(number).replace('[', '').rstrip(']')

            create_username = str(get_nienkhoa.ten_nienkhoa) + \
                str(get_hedt.key) + str(get_nganh.key_nganh) + new_string

            # Lấy password làm ngày sinh
            password = request.POST.get('ngaysinh')

            # Tiến hành tạo mới user
            new_user = User.objects.create_user(
                username=create_username, email=None, password=password)

            u = User.objects.get(username=new_user.username)
            # get data sinhvien
            hoten = request.POST.get('hoten')
            ngaysinh = request.POST.get('ngaysinh')
            print('SINNHHH: ', ngaysinh)
            gender = request.POST.get('gender')
            nganh = request.POST.get('nganh')
            hedt = request.POST.get('hedt')
            nien_khoa = request.POST.get('nien_khoa')
            nganh_name = Nganh.objects.get(id=int(nganh))
            he_dt_name = HeDT.objects.get(id=int(hedt))
            nien_khoa_name = NienKhoa.objects.get(id=int(nien_khoa))

            check = True
            try:
                SinhVien.objects.create(
                    user=u,
                    hoten=hoten,
                    ngaysinh=ngaysinh,
                    gender=gender,
                    nganh=nganh_name,
                    hedt=he_dt_name,
                    nien_khoa=nien_khoa_name
                )
            except:
                check = False
            if check:
                messages.add_message(
                    request, messages.SUCCESS, 'Thêm sinh viên thành công !')
            else:
                messages.add_message(
                    request, messages.ERROR, 'Thêm sinh viên thất bại !')

        except:
            messages.add_message(request, messages.WARNING,
                                 'Các trường không được phép trống !')

    context = {
        'form': form
    }
    return render(request, 'create_form/create.html', context)


def import_csv(request):
    if not request.user.is_superuser:
        raise PermissionDenied
    if request.method == 'GET':
        return render(request, 'create.html')
    else:
        file_name = request.FILES.get('file_excel')

        # GET name value:
        khoa_id = request.POST.get('nien_khoa')
        nganh_id = request.POST.get('nganh')
        hedt_id = request.POST.get('hedt')

        get_nganh = Nganh.objects.get(pk=nganh_id)
        get_hedt = HeDT.objects.get(pk=hedt_id)
        get_nienkhoa = NienKhoa.objects.get(pk=khoa_id)

        wb = load_workbook(file_name)
        sheet = wb.active
        max_col = sheet.max_column
        print("MAX_COL:", max_col)
        max_row = sheet.max_row
        arr = []
        for i in range(1, max_row+1):  # i = 1
            for j in range(1, max_col+1):  # j = 1
                # hang 1 cot 1, hang 1 cot 2, hang 1 cot 3
                c = sheet.cell(row=i, column=j)
                print(i, j)
                print(c.value)

                # Chuyển ngày tháng về đúng định dạng
                if isinstance(c.value, datetime.datetime):
                    data = c.value
                    data_datetime = data.strftime("%Y-%m-%d")
                    c.value = data_datetime
                arr.append(c.value)
        new_arr = [arr[i * max_col:(i + 1) * max_col]
                   for i in range((len(arr) + max_col - 1) // max_col)]
        run = True
        for i, a in enumerate(new_arr):
            number = random.sample(range(4000), 1)
            # Trả về 1 array với 1 phần tử
            new_string = str(number).replace('[', '').rstrip(']')
            create_username = str(get_nienkhoa.ten_nienkhoa) + \
                str(get_hedt.key) + str(get_nganh.key_nganh) + new_string
            if not new_arr[i][1]:
                messages.add_message(
                    request, messages.WARNING, "Các trường không được bỏ trống: Họ tên, ngày sinh.")
            else:
                print(new_arr[i])
                try:
                    new_user = User.objects.create_user(
                        username=create_username,
                        email=None,
                        password=new_arr[i][1]
                    )
                    u = User.objects.get(username=new_user.username)
                    print(u)
                    SinhVien.objects.create(
                        user=u,
                        hoten=new_arr[i][0],
                        ngaysinh=new_arr[i][1],
                        gender=new_arr[i][2],
                        address=new_arr[i][3],
                        nganh=Nganh.objects.get(id=nganh_id),
                        hedt=HeDT.objects.get(id=hedt_id),
                        nien_khoa=NienKhoa.objects.get(id=khoa_id)
                    )
                except:
                    print("DM")
                    User.objects.get(username=create_username).delete()
    if run:
        messages.add_message(request, messages.SUCCESS,
                             'Thêm dữ liệu thành công')
    else:
        messages.add_message(request, messages.ERROR, 'Thêm dữ liệu thất bại.')
    return redirect('list_student')


def list_student_lop(request, id):
    lop_id = Lop.objects.get(pk=id)
    list_sv = SinhVien.objects.filter(lop=lop_id)
    count_sv = list_sv.count()
    request.session['count'] = count_sv
    context = {
        'lop_id': lop_id,
        'count_sv': count_sv,
        'list_sv': list_sv
    }
    request.session['student'] = lop_id.id
    return render(request, 'list_student_lop.html', context)


def detail_student(request, id):
    if not request.user.is_superuser:
        raise PermissionDenied
    sv = SinhVien.objects.get(pk=id)
    context = {
        'sv': sv
    }
    return render(request, 'detail_student.html', context)


def delete_student(request, id):
    if not request.user.is_superuser:
        raise PermissionDenied
    id_lop = request.session['student']
    l = Lop.objects.get(id=id_lop)
    sv = SinhVien.objects.get(pk=id)
    sv.lop.remove(l)
    messages.add_message(request, messages.SUCCESS, 'Thao tác thành công!')
    return redirect('list_student_lop', id_lop)


def search(request):
    if not request.user.is_superuser:
        raise PermissionDenied
    key = request.GET.get('search')
    result_masv = SinhVien.objects.filter(user__username__icontains=key)
    result_hoten = SinhVien.objects.filter(hoten__icontains=key)
    result_lop = Lop.objects.filter(ma_lop__icontains=key)
    context = {
        'result_masv': result_masv,
        'result_hoten': result_hoten,
        'result_lop': result_lop,
        'key': key
    }
    return render(request, 'search.html', context)


def list_student(request):
    if not request.user.is_superuser:
        raise PermissionDenied
    data = SinhVien.objects.all()
    context = {
        'data': data
    }
    return render(request, 'list_student.html', context)


def delete_user(request, id):
    if not request.user.is_superuser:
        raise PermissionDenied
    user_id = SinhVien.objects.get(pk=id)
    user_id.delete()
    messages.add_message(request, messages.ERROR, 'Đã xóa sinh viên !')
    return redirect('list_student')


# CREATE FORM
def create_lop(request):
    if not request.user.is_superuser:
        raise PermissionDenied
    form = CreateLop()
    if request.method == 'POST':
        form = CreateLop(request.POST)
        if form.is_valid():
            form.cleaned_data
            form.save()
            messages.add_message(request, messages.SUCCESS,
                                 'Thêm lớp thành công!')
            return redirect('lop')
    context = {
        'form': form
    }
    return render(request, 'create_form/create_lop.html', context)


def create_teacher(request):
    if not request.user.is_superuser:
        raise PermissionDenied
    form = CreateGiangVien()
    if request.method == 'POST':
        form = CreateGiangVien(request.POST)
        if form.is_valid():
            form.cleaned_data
            form.save()
            user_id = request.POST.get('ma_gv')
            email = request.POST.get('email')
            ngaysinh = request.POST.get('ngaysinh')
            u = User.objects.create_user(
                username=user_id,
                email=email,
                password=str(ngaysinh)
            )
            u.is_staff = True
            u.save()
            gv = GiaoVien.objects.get(ma_gv=user_id)
            gv.user = u
            gv.save()
            messages.add_message(request, messages.SUCCESS,
                                 'Thêm giảng viên thành công!')
            return redirect('list_teacher')

    context = {
        'form': form
    }
    return render(request, 'create_form/create_teacher.html', context)


def list_teacher(request):
    if not request.user.is_superuser:
        raise PermissionDenied
    data = GiaoVien.objects.all()
    context = {
        'data': data
    }
    return render(request, 'list_teacher.html', context)


def delete_teacher(request, id):
    if not request.user.is_superuser:
        raise PermissionDenied
    gv = GiaoVien.objects.get(pk=id)
    gv.delete()
    messages.add_message(request, messages.ERROR, 'Đã xóa giáo viên')
    return redirect('list_teacher')


def detail_teacher(request, id):
    if not request.user.is_superuser:
        raise PermissionDenied
    gv = GiaoVien.objects.get(pk=id)

    context = {
        'gv': gv
    }
    return render(request, 'detail_teacher.html', context)


def delete_lop(request, id):
    if not request.user.is_superuser:
        raise PermissionDenied
    l = Lop.objects.get(pk=id)
    l.delete()
    return redirect('lop')


def lop(request):
    if not request.user.is_superuser:
        raise PermissionDenied
    form = FormHeDT()
    if request.method == 'POST':
        form = FormHeDT(request.POST)
        he_dt = request.POST.get('he_dt')
        if form.is_valid():
            form.cleaned_data
        if he_dt == '':
            he_dt = 1
        he_id = HeDT.objects.get(pk=he_dt)
        data = Lop.objects.filter(he_dt=he_dt)
    else:
        data = Lop.objects.all()
    return render(request, 'lop.html', context={
        'form': form,
        'data': data
    })


def hoc_phi(request):
    if not request.user.is_superuser:
        raise PermissionDenied
    user = User.objects.get(username=request.user)
    sv = SinhVien.objects.get(user=user)
    found = False
    input_msv = request.POST.get('masv', request.session.get('get_msv'))
    if input_msv is None:
        input_msv = ""
    msv = input_msv.upper()
    request.session['get_msv'] = msv
    if request.method == 'POST':
        find_user = User.objects.filter(username=msv)
        if len(find_user) == 0:
            messages.add_message(request, messages.ERROR,
                                 'Không tìm thấy dữ liệu')

    data_sv = SinhVien.objects.filter(user__username=msv)
    try:
        data = HocPhiConNo.objects.filter(user__username=msv)
        s = SinhVien.objects.get(user__username=msv)
        form = DongHocPhi(instance=s)
        found = True
        da_dong = HocPhiDaDong.objects.filter(user__username=msv)
        total_no = []
        total_du = []
        for i in data:
            if i.so_tien_con_no <= 0:
                i.so_du += i.so_tien_con_no * -1
                i.so_tien_con_no = 0
                i.save()

            total_no.append(i.so_tien_con_no)
            total_du.append(i.so_du)
        if sum(total_no) > sum(total_du):
            result = sum(total_no) - sum(total_du)
        else:
            result = 0
        sum_da_dong = []
        for i in da_dong:
            sum_da_dong.append(i.so_tien)
    except:
        data = None,
        s = None
        form = None
        da_dong = ''
        result = ''

    return render(request, 'hocphi.html', context={
        'data': data,
        'found': found,
        'sv': s,
        'form': form,
        'da_dong': da_dong,
        'result': result,
    })


def dong_hoc_phi(request):
    if not request.user.is_superuser:
        raise PermissionDenied
    form = DongHocPhi()
    receive = request.session['get_msv']
    if request.method == 'POST':
        form = DongHocPhi(request.POST)
        hoc_ki = request.POST.get('hoc_ki')
        user = request.POST.get('user')
        u = User.objects.get(id=user)
        hk = HocKi.objects.get(ten_hk=hoc_ki)
        so_tien = request.POST.get('so_tien')
        if form.is_valid():
            form.cleaned_data
            form.save()
            History.objects.create(
                user=u,
                actions=f'đã đóng học phí : {so_tien} VNĐ'
            )
            data = HocPhiDaDong.objects.filter(user=u, hoc_ki=hoc_ki)
            arr = []
            for i in data:
                arr.append(i.so_tien)
            no = HocPhiConNo.objects.get(user=u, hoc_ki=hoc_ki)
            no.so_tien_con_no -= arr[-1]
            no.save()

            messages.add_message(request, messages.SUCCESS,
                                 'Thao tác thành công')

    return redirect('hoc_phi')


def delete_gd(request, id):
    if not request.user.is_superuser:
        raise PermissionDenied
    get_msv = request.session['get_msv']
    data = HocPhiDaDong.objects.get(pk=id)
    tien_no = HocPhiConNo.objects.get(
        user__username=get_msv, hoc_ki=data.hoc_ki)
    if tien_no.so_du >= 0:
        tien_no.so_du -= data.so_tien
        tien_no.save()
        if tien_no.so_du < 0:
            tien_no.so_tien_con_no += tien_no.so_du * -1
            tien_no.so_du = 0
            tien_no.save()
    data.delete()
    History.objects.create(
        user=request.user,
        actions=f'xoá giao dịch {data.user} số tiền {data.so_tien} VNĐ'
    )
    messages.add_message(request, messages.SUCCESS, 'Thao tác thành công!')
    return redirect('hoc_phi')


def menu_add(request):
    return render(request, 'menu_add.html')


def history(request):
    if not request.user.is_superuser:
        raise PermissionDenied
    data = History.objects.all()
    return render(request, 'history.html', context={
        'data': data
    })


def send_messages(request):
    sv = SinhVien.objects.all()
    if request.method == 'POST':
        message = request.POST.get('messages')
        list_user = request.POST.getlist('receiver')
        title = request.POST.get('title')
        if not list_user:
            messages.add_message(request, messages.ERROR,
                                 "Chưa chọn người nhận.")
        else:
            for user in list_user:
                s = SinhVien.objects.get(pk=user)
                print("USER: ", s)
                Messages.objects.create(
                    title=title,
                    sender=request.user,
                    messages=message,
                    receiver=s,
                )
            messages.add_message(request, messages.SUCCESS,
                                 'Gửi tin nhắn thành công!')
    form = NganhForm()
    return render(request, 'send_messages.html', context={
        'sv': sv,
        'form': form
    })


def filter_student(request):
    khoa = request.POST.get('khoa')
    form = NganhForm()
    if request.method == 'POST':
        form = NganhForm(request.POST)
        if form.is_valid():
            form.cleaned_data
        sv = SinhVien.objects.filter(nganh__id=khoa)
        print(sv)
    return render(request, 'send_messages.html', context={
        'sv': sv,
        'form': form
    })
